#xlsx   xls
#openpyxl ->  xlsx
from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    dic = {}#字典


    #打开文件
    file = load_workbook(filename= path)
    #所有表格的名称
    #print(file.get_sheet_names())
    sheets = file.get_sheet_names()
    print(len(sheets))

    #拿出一个表格
    #sheet = file.get_sheet_by_name(sheets[0])
    #最大行数
    #print(sheet.max_row)
    #最大列数
    #print(sheet.max_column)
    #表名
    #print(sheet.title)


    for sheetName in sheets:
        sheet = file.get_sheet_by_name(sheetName)
        #一张表所有数据
        sheetInfor = []
        for lineNum in range(1, sheet.max_row + 1):
            lineList = []
            for columnNum in range(1, sheet.max_column + 1):
                value = sheet.cell(row=lineNum, column=columnNum).value
                if value != None:     #可根据实际情况确认是否需要取出None
                    lineList.append(value)
            sheetInfor.append(lineList)

        #将一张表的数据存到字典
        dic[sheetName] = sheetInfor
    return dic




#不能处理xls文件
path = r"C:\Users\Zhangyadi\Desktop\project\115word自动化办公\sunck.xlsx"
dic = readXlsxFile(path)
print(dic)
print(dic["Sheet1"])#只打印一张表
print(len(dic))