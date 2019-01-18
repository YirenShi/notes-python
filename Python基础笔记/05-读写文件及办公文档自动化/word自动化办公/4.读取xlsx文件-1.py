import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Zhangyadi\Desktop\project\115word自动化办公\sunck.xlsx")

# 获取workbook中所有的表格
sheets = wb.get_sheet_names()

print(sheets)

# 循环遍历所有sheet
for i in range(len(sheets)):
    sheet = wb.get_sheet_by_name(sheets[i])

    print('\n\n第' + str(i + 1) + '个sheet: ' + sheet.title + '->>>')



from openpyxl import load_workbook
wb = load_workbook(r"C:\Users\Zhangyadi\Desktop\project\115word自动化办公\sunck.xlsx")
sheets = wb.get_sheet_names()
for i,sheet in enumerate(sheets,1):
    print(i,sheet)
